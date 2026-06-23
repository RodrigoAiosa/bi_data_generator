"""generators/dicionario.py — Gera dicionário de dados em Excel."""

import io
import pandas as pd

# Descrições por padrão de nome de coluna
_DESC: dict[str, str] = {
    # IDs
    "id_":          "Identificador único",
    "sk_":          "Surrogate key (chave substituta)",
    # Datas
    "data_":        "Data do evento",
    "id_data":      "Chave para dCalendario",
    # Valores
    "valor_":       "Valor monetário (R$)",
    "receita":      "Receita gerada (R$)",
    "custo_":       "Custo associado (R$)",
    "preco_":       "Preço unitário (R$)",
    "desconto":     "Desconto aplicado",
    "margem":       "Margem percentual",
    # Quantidades
    "qtd_":         "Quantidade",
    "volume_":      "Volume físico",
    "quantidade":   "Quantidade de unidades",
    # Status / flags
    "status":       "Status do registro",
    "ativo":        "Indica se o registro está ativo",
    "cancelado":    "Indica se foi cancelado",
    "aprovado":     "Indica se foi aprovado",
    # Textos comuns
    "nome":         "Nome descritivo",
    "descricao":    "Descrição detalhada",
    "tipo_":        "Classificação por tipo",
    "categoria":    "Categoria de negócio",
    "canal":        "Canal de origem ou venda",
    "regiao":       "Região geográfica",
    "uf":           "Unidade federativa (estado)",
    "cidade":       "Município",
    "cnpj":         "CNPJ da empresa",
    "cpf":          "CPF do indivíduo",
    "email":        "Endereço de e-mail",
    "telefone":     "Número de telefone",
    # Métricas
    "pct":          "Percentual (%)",
    "taxa_":        "Taxa percentual (%)",
    "score":        "Pontuação / índice",
    "avaliacao":    "Nota de avaliação",
    "nps":          "Net Promoter Score",
    "mrr":          "Monthly Recurring Revenue",
    "arr":          "Annual Recurring Revenue",
    "churn":        "Indicador de cancelamento",
}

_TIPO_MAP = {
    "int64":    "Inteiro",
    "int32":    "Inteiro",
    "float64":  "Decimal",
    "float32":  "Decimal",
    "object":   "Texto",
    "bool":     "Booleano",
    "datetime64[ns]": "Data/Hora",
}

_TABELA_DESC = {
    "FatoVenda":        "Registros de vendas realizadas",
    "FatoPedido":       "Pedidos de clientes",
    "FatoTransacao":    "Transações financeiras",
    "FatoAtividade":    "Atividades e interações registradas",
    "FatoOportunidade": "Oportunidades no funil de vendas",
    "FatoProducao":     "Registros de produção",
    "FatoPartida":      "Partidas e eventos esportivos",
    "FatoReserva":      "Reservas e hospedagens",
    "FatoPlay":         "Reproduções de conteúdo (plays)",
    "FatoAssinatura":   "Assinaturas e contratos recorrentes",
    "FatoProcesso":     "Processos jurídicos",
    "FatoExtracao":     "Extrações minerais",
    "FatoViagem":       "Corridas e viagens",
    "FatoHorasTrabalhadas": "Horas trabalhadas por colaborador",
    "FatoDespesa":      "Despesas orçamentárias",
    "FatoReceita":      "Arrecadações e receitas governamentais",
    "FatoLicitacao":    "Licitações e contratos públicos",
    "FatoEstoque":      "Movimentações de estoque",
    "FatoCusto":        "Custos operacionais",
    "dCalendario":      "Tabela calendário para análises temporais",
}


def _inferir_desc(col: str) -> str:
    col_lower = col.lower()
    for pattern, desc in _DESC.items():
        if col_lower.startswith(pattern) or col_lower == pattern or pattern in col_lower:
            return desc
    return "—"


def gerar_dicionario(nome_setor: str, tabelas: dict[str, pd.DataFrame]) -> bytes:
    """
    Gera um arquivo Excel com o dicionário de dados de todas as tabelas.

    Retorna bytes prontos para download.
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ── Aba Resumo ────────────────────────────────────────────────────
        resumo_rows = []
        for tname, tdf in tabelas.items():
            fato_flag = "✅ Fato" if tname.startswith("Fato") else (
                "📅 Calendário" if tname.startswith("dCal") else "📋 Dimensão"
            )
            resumo_rows.append({
                "Tabela":     tname,
                "Tipo":       fato_flag,
                "Linhas":     len(tdf),
                "Colunas":    len(tdf.columns),
                "Descrição":  _TABELA_DESC.get(tname, f"Tabela do setor {nome_setor}"),
            })

        df_resumo = pd.DataFrame(resumo_rows)
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        _format_sheet(writer, "Resumo", df_resumo)

        # ── Aba por tabela ────────────────────────────────────────────────
        for tname, tdf in tabelas.items():
            rows = []
            for col in tdf.columns:
                dtype = str(tdf[col].dtype)
                sample = tdf[col].dropna().iloc[0] if len(tdf[col].dropna()) > 0 else "—"
                rows.append({
                    "Coluna":     col,
                    "Tipo":       _TIPO_MAP.get(dtype, dtype),
                    "Descrição":  _inferir_desc(col),
                    "Exemplo":    str(sample)[:60],
                    "Nulos":      int(tdf[col].isna().sum()),
                    "Únicos":     int(tdf[col].nunique()),
                })

            df_dict = pd.DataFrame(rows)
            sheet = tname[:31]  # Excel max 31 chars
            df_dict.to_excel(writer, sheet_name=sheet, index=False)
            _format_sheet(writer, sheet, df_dict)

    return output.getvalue()


def _format_sheet(writer, sheet_name: str, df: pd.DataFrame) -> None:
    """Aplica formatação básica: largura de colunas e header bold."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        ws = writer.sheets[sheet_name]

        # Header
        header_fill = PatternFill("solid", fgColor="1E1B4B")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Largura automática
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)
    except Exception:
        pass  # Formatação é opcional
